import openpyxl
from openpyxl import *


class AbrirExcel:

    def __init__(self, driver):
        self.driver = driver

    # Funcion que obtiene el numero de filas que tenga en archivo Excel
    def obtenerNumeroFilas(self, ruta, nombreHoja):
        WorkBook = openpyxl.load_workbook(ruta)
        sheet = WorkBook[nombreHoja]
        return sheet.max_row

    def obtenerNumeroColumnas(self, file,  nombreHoja):
        WorkBook = openpyxl.load_workbook(file)
        sheet = WorkBook[nombreHoja]
        return sheet.max_column

    def leerDatosDeExcel(self, ruta, nombreHoja, fila_num, columna_num):
        WorkBook = openpyxl.load_workbook(ruta)
        sheet = WorkBook[nombreHoja]
        return sheet.cell(row=fila_num, column=columna_num).value

    def esribirDatosEnExcel(self, ruta, nombreHoja, fila_num, columna_num, dato):
        WorkBook = openpyxl.load_workbook(ruta)
        sheet = WorkBook[nombreHoja]
        sheet.cell(row=fila_num, column=columna_num).value = dato
        WorkBook.save(ruta)

    def crearNuevaLibro(self, ruta, nombreArchivo):
        # Instanciamos un objeto de la clase WorkBook()
        libro = Workbook()
        # Creamos una variable en la cual guardamos le ruta donde ira el archivo concatenado con el nombre del mismo
        filesheet = ruta+"//"+nombreArchivo
        # Mediante el medoto save(), guardamos dicho libro de excel
        libro.save(filesheet)

    def insertarDatosEnHojaExcel(self):
        pass




